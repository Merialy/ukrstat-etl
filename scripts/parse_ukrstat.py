
""" Укрстат - JSON/CSV конвертер | Проєкт: «Дані» """

import pandas as pd
import json
import csv
import os
import sys
import logging
from datetime import datetime
from pathlib import Path

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ------ Конфігурація ------
CONFIG = {
    "skip_rows": 3, # Рядків-заголовків зверху у типовому файлі Укрстату
    "na_values": ["-", "–", "...", "н/д", "х", ""],
    "encoding_csv": "utf-8-sig", # utf-8-sig для сумісності з Excel
}

# ------ Утиліти ------
def detect_header_row(df_raw: pd.DataFrame) -> int:
    """Знаходить перший рядок, де більшість клітинок заповнені (реальний заголовок)."""
    for i, row in df_raw.iterrows():
        filled = row.notna().sum()
        if filled >= max(3, len(row) * 0.4):
            return i
    return 0


def clean_column_name(name: str) -> str:
    """Нормалізує назву стовпця: прибирає пробіли, переноси, лапки."""
    if not isinstance(name, str):
        name = str(name)
    return ( name.strip().replace("\n", " ")
                         .replace("\r", "")
                         .replace('"', "")
                         .replace("'", "")
    )


def infer_types(df: pd.DataFrame) -> pd.DataFrame:
    """Намагається конвертувати числові стовпці."""
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col])
        except (ValueError, TypeError):
            pass
    return df


# ------ Основна логіка ------
def load_excel(filepath: str, sheet_name: int | str = 0) -> pd.DataFrame:
    """Завантажує Excel-файл Укрстату, автоматично визначаючи рядок заголовку."""
    logger.info(f"Читаю файл: {filepath}")

    # Перше сканування — щоб знайти заголовок
    df_raw = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    header_row = detect_header_row(df_raw)
    logger.info(f"Заголовок знайдено у рядку: {header_row}")

    df = pd.read_excel(
        filepath,
        sheet_name=sheet_name,
        header=header_row,
        na_values=CONFIG["na_values"],
    )

    # Прибираємо повністю порожні рядки / стовпці
    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)

    # Нормалізуємо назви стовпців
    df.columns = [clean_column_name(c) for c in df.columns]

    df = infer_types(df)
    logger.info(f"Завантажено {len(df)} рядків, {len(df.columns)} стовпців")
    return df


def add_metadata(df: pd.DataFrame, source_file: str) -> list[dict]:
    """Перетворює DataFrame на список словників із метаданими."""
    records = df.to_dict(orient="records")
    meta = {
        "_source": Path(source_file).name,
        "_parsed_at": datetime.utcnow().isoformat() + "Z",
        "_total_rows": len(records),
    }
    return records, meta


def save_json(records: list[dict], meta: dict, output_path: str) -> None:
    """Зберігає дані у JSON."""
    payload = {"metadata": meta, "data": records}
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    logger.info(f"JSON збережено: {output_path}")


def save_csv(records: list[dict], output_path: str) -> None:
    """Зберігає дані у CSV (utf-8-sig для Excel-сумісності)."""
    if not records:
        logger.warning("Немає даних для збереження у CSV")
        return
    # Замінюємо float('nan') на порожній рядок перед записом
    import math
    cleaned = [
        {k: ("" if isinstance(v, float) and math.isnan(v) else v) for k, v in row.items()}
        for row in records
    ]
    with open(output_path, "w", encoding=CONFIG["encoding_csv"], newline="") as f:
        writer = csv.DictWriter(f, fieldnames=cleaned[0].keys())
        writer.writeheader()
        writer.writerows(cleaned)
    logger.info(f"CSV збережено: {output_path}")

# ------ Генерація тестового Excel ------
def generate_sample_excel(filepath: str) -> None:
    """Генерує демо-файл у стилі Укрстату для тестування скрипту."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Імітація шапки Укрстату
    ws["A1"] = "Державна служба статистики України"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Таблиця 1. Середня заробітна плата за регіонами (грн)"
    ws["A3"] = "(за даними підприємств)"

    headers = ["Регіон", "2020", "2021", "2022", "2023", "2024"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    data = [
        ["Вінницька", 10234, 12456, 14100, 16890, 19200],
        ["Дніпропетровська", 12000, 14500, 16200, 19500, 22100],
        ["Івано-Франківська", 9800, 11900, 13700, 16200, 18500],
        ["Київська", 15600, 18900, 21400, 25600, 29800],
        ["Львівська", 11200, 13600, 15500, 18700, 21300],
        ["Харківська", 11800, 14200, "-", 15100, 18400],  # '-' — відсутні дані
        ["Одеська", 10900, 13100, 14900, 17800, 20200],
        ["м. Київ", 22000, 26500, 30100, 36200, 42000],
    ]
    for row in data:
        ws.append(row)

    ws.column_dimensions["A"].width = 22
    wb.save(filepath)
    logger.info(f"Демо-файл створено: {filepath}")

# ------ Точка входу ------
def main():
    base_dir = Path(__file__).parent.parent
    sample_excel = base_dir / "data" / "sample_ukrstat.xlsx"
    output_json  = base_dir / "data" / "output.json"
    output_csv   = base_dir / "data" / "output.csv"

    sample_excel.parent.mkdir(parents=True, exist_ok=True)

    # Якщо файл не передано аргументом — генеруємо демо
    input_file = sys.argv[1] if len(sys.argv) > 1 else str(sample_excel)

    if not Path(input_file).exists():
        logger.info("Реального файлу не знайдено — генерую демо-дані...")
        generate_sample_excel(input_file)

    df = load_excel(input_file)
    records, meta = add_metadata(df, input_file)

    save_json(records, meta, str(output_json))
    save_csv(records, str(output_csv))

    # Короткий звіт
    print("\n" + "="*50)
    print(f"✅ Оброблено рядків: {meta['_total_rows']}")
    print(f"✅ JSON:             {output_json}")
    print(f"✅ CSV:              {output_csv}")
    print("="*50)

if __name__ == "__main__":
    main()
